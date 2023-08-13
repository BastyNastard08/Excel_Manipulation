import openpyxl
def column_finder():
    wb = openpyxl.load_workbook("Book1.xlsx")
    ws = wb["mark"]
    titles = ws.iter_cols(min_col=1,max_col=3,min_row=1,max_row=1)
    for title in titles:
        for cell in title:
            if cell.value.lower() == "address":
                return cell.coordinate[0]
numbers = []
def number_adder():
    global numbers
    wb = openpyxl.load_workbook("Book1.xlsx")
    ws = wb["mark"]
    maxrow = ws.max_row
    rows = ws[column_finder()+"2":f"{column_finder()}{maxrow}"]
    rowcount = 2
    for row in rows:
        for cell in row:
            coordinate11 = f"{chr(ord(column_finder())+1)}{rowcount}"
            coordinate10 = f"{chr(ord(column_finder())+2)}{rowcount}"
            rowcount += 1
            cellval = f"{cell.value}"
            for i in range(len(cellval)):
                if cellval[i].isnumeric():
                    try:
                        if cellval[i].isnumeric():
                            if cellval[i+10].isnumeric() and cellval[i:i+10].isnumeric():
                                numbers.append(cellval[i:i+11])
                                i += 9
                                ws[coordinate11] = cellval[i:i+11]
                                wb.save("Book1.xlsx")
                                continue
                    except IndexError:
                        pass
                    try:
                        if cellval[i+9].isnumeric() and cellval[i:i+9].isnumeric():
                            numbers.append(cellval[i:i+10])
                            i += 9
                            ws[coordinate10] = cellval[i:i+10]
                            wb.save("Book1.xlsx")
                    except IndexError:
                        continue
print(number_adder())
