import openpyxl

def column_finder(ws):
    titles = ws[1]
    for index, cell in enumerate(titles, start=1):
        if cell.value.lower() == "address":
            return openpyxl.utils.get_column_letter(index)

def number_adder(ws, column_letter):
    numbers = []
    maxrow = ws.max_row
    column_range = ws[f"{column_letter}2:{column_letter}{maxrow}"]
    for row in column_range:
        for cell in row:
            cellval = str(cell.value)
            for i in range(len(cellval)):
                if cellval[i].isnumeric():
                    try:
                        if cellval[i+10].isnumeric() or cellval[i+8].isnumeric():
                            if cellval[i:i+10].isnumeric():
                                numbers.append(cellval[i:i+10])
                            elif cellval[i:i+8].isnumeric():
                                numbers.append(cellval[i:i+8])
                    except IndexError:
                        break
    return numbers

if __name__ == "__main__":
    wb = openpyxl.load_workbook("Book1.xlsx")
    ws = wb["mark"]
    col_letter = column_finder(ws)
    result = number_adder(ws, col_letter)
    print(result)